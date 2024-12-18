import openpyxl as xl
from openpyxl.chart import BarChart, Reference #for the chart

#updates the spreadsheet by changing price in 3rd column to 90% of it
def process_workbook(filename):
    # path is in the current directory
    wb = xl.load_workbook(filename)  # loads workbook, returns workbook object
    # probably a map
    sheet = wb['Sheet1']  # gets the sheet
    PRICE_COLUMN = 3

    # cell = sheet['A1'] #gets cell A1
    # cell = sheet.cell(1, 1) #row 1 column 1
    # print(cell.value) value of the cell
    # print(sheet.max_row) amount of rows

    # applying 90% to the price
    for row in range(2, sheet.max_row + 1):  # row 1 is the heading
        cell = sheet.cell(row, PRICE_COLUMN)
        corrected_price = cell.value * 0.9
        # adding new column for new prices
        corrected_price_cell = sheet.cell(row,
                                          PRICE_COLUMN + 1)  # gets the cell of the column after
        # the price
        corrected_price_cell.value = corrected_price  # changes cell value to corrected price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)  # the sheet, get row 2-4, column 4

    chart = BarChart()
    chart.add_data(values)  # add values to chart
    sheet.add_chart(chart, 'e2')  # add chart top left corner there
    wb.save(filename)

process_workbook('transactions.xlsx')